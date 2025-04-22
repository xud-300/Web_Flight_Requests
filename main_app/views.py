import json
from django.shortcuts import get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, View
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden, HttpResponseBadRequest
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django import forms
from django.template.loader import render_to_string
from .models import FlightRequest, RequestHistory, Object, ObjectType, TelegramUser, UndoAction
from main_app.models import User
from .forms import FlightRequestCreateForm, FlightRequestEditForm
from django.views.decorators.http import require_POST, require_GET
from django.utils.decorators import method_decorator
from .models import FlightRequest, FlightResultFile
from .upload_forms import OrthoUploadForm, LaserUploadForm, PanoramaUploadForm, OverviewUploadForm
from django.utils.dateparse import parse_datetime


# Список заявок
class FlightRequestListView(LoginRequiredMixin, ListView):
    model = FlightRequest
    template_name = 'main_app/requests_list.html'
    context_object_name = 'requests'
    paginate_by = 25

    def get_queryset(self):
        qs = FlightRequest.objects.all()

        # Фильтрация по GET-параметрам
        status = self.request.GET.get('status')
        object_type = self.request.GET.get('object_type')
        object_name = self.request.GET.get('object_name')
        shooting_type = self.request.GET.get('shooting_type')
        shoot_date_from = self.request.GET.get('shoot_date_from')
        shoot_date_to = self.request.GET.get('shoot_date_to')
        
        if status:
            qs = qs.filter(status=status)
        
        if object_type:
            qs = qs.filter(object_type_id=object_type)
        
        if object_name:
            qs = qs.filter(object_name_id=object_name)
        
        if shooting_type:
            filter_kwargs = {shooting_type: True}
            qs = qs.filter(**filter_kwargs)
        
        # Фильтрация по диапазону дат съемки (интервалы пересекаются, если shoot_date_from(record) <= shoot_date_to(filter)
        # и shoot_date_to(record) >= shoot_date_from(filter))
        if shoot_date_from and shoot_date_to:
            qs = qs.filter(shoot_date_from__lte=shoot_date_to, shoot_date_to__gte=shoot_date_from)
        elif shoot_date_from:
            qs = qs.filter(shoot_date_to__gte=shoot_date_from)
        elif shoot_date_to:
            qs = qs.filter(shoot_date_from__lte=shoot_date_to)
        
        # Сортировка: получаем GET-параметры сортировки
        sort_field = self.request.GET.get('sort')
        order = self.request.GET.get('order', 'asc')
        
        if sort_field == 'shoot_date':
            if order == 'desc':
                qs = qs.order_by('-shoot_date_to')
            else:
                qs = qs.order_by('shoot_date_from')
        else:
            allowed_sort_fields = {
                'status': 'status',
                'id': 'id',
                'object_type': 'object_type__type_name'
            }
            if sort_field in allowed_sort_fields:
                field_name = allowed_sort_fields[sort_field]
                if order == 'desc':
                    qs = qs.order_by('-' + field_name)
                else:
                    qs = qs.order_by(field_name)
            else:
                qs = qs.order_by('-id')
        
        return qs


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import ObjectType
        context['object_types'] = ObjectType.objects.all()

        # Копируем GET-параметры и удаляем параметр "page", чтобы избежать дублирования при формировании ссылок пагинации
        params = self.request.GET.copy()
        if 'page' in params:
            del params['page']
        context['querystring'] = params.urlencode()

        return context



# Создание заявки через модальное окно.
class FlightRequestCreateView(LoginRequiredMixin, CreateView):
    model = FlightRequest
    form_class = FlightRequestCreateForm
    template_name = 'modal_create.html'
    success_url = reverse_lazy('requests_list')

    def form_valid(self, form):
        # Привязываем пользователя из TelegramUser или создаем нового
        try:
            user_obj = TelegramUser.objects.get(id=self.request.user.id)
        except TelegramUser.DoesNotExist:
            user_obj = TelegramUser.objects.create(
                id=self.request.user.id,
                telegram_id=getattr(self.request.user, 'telegram_id', 0),
                username=self.request.user.username or '',
                full_name=getattr(self.request.user, 'full_name', '')
            )
        form.instance.user = user_obj
        form.instance.username = self.request.user.profile.full_name

        # Сохраняем сам запрос и историю
        self.object = form.save()
        RequestHistory.objects.create(
            flight_request=self.object,
            changed_by=self.request.user,
            changes=json.dumps({"created": ["", "Заявка создана"]}, ensure_ascii=False)
        )

        # Если AJAX — вернем JSON, иначе стандартный redirect
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'redirect_url': str(self.success_url)})
        return super().form_valid(form)

    def form_invalid(self, form):
        # При AJAX — возвращаем JSON со всеми ошибками (полевыми и неполевыми)
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'errors': form.errors})
        return super().form_invalid(form)


# Редактирование заявки через модальное окно.
class FlightRequestUpdateView(LoginRequiredMixin, UpdateView):
    model = FlightRequest
    form_class = FlightRequestEditForm
    template_name = 'main_app/modal_edit.html'
    success_url = reverse_lazy('requests_list')

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        # только админы или автор могут редактировать
        if not (request.user.is_staff or (hasattr(request.user, 'profile') and request.user.profile.role == 'admin')):
            if self.object.user_id != request.user.id:
                return HttpResponseForbidden("У вас нет прав для редактирования этой заявки.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object_types'] = ObjectType.objects.all()
        context['object_names'] = Object.objects.filter(object_type=self.object.object_type)
        context['edit_url'] = reverse_lazy('request_edit', args=[self.object.id])
        context['is_editable'] = (
            self.object.status != "Выполнена"
            or self.request.user.is_staff
            or (hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'admin')
        )
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data(object=self.object, form=self.get_form())
        # при AJAX запросе возвращаем только тело модалки
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('ajax'):
            html = render_to_string("main_app/modal_edit.html", {**context, 'ajax': True}, request=request)
            return HttpResponse(html)
        return super().get(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # для админа показываем выбор статуса
        if self.request.user.is_staff or (hasattr(self.request.user, 'profile') and self.request.user.profile.role == 'admin'):
            form.fields['status'] = forms.ChoiceField(
                choices=[('Новая','Новая'),('В работе','В работе'),('Выполнена','Выполнена')],
                initial=self.object.status,
                widget=forms.Select(attrs={'class':'form-control','id':'id_edit_status'})
            )
        return form

    def form_valid(self, form):
        response = super().form_valid(form)
        # сохраняем историю изменений
        changes = {}
        for field in form.changed_data:
            old = form.initial.get(field)
            new = form.cleaned_data.get(field)
            changes[field] = [str(old), str(new)]
        RequestHistory.objects.create(
            flight_request=self.object,
            changed_by=self.request.user,
            changes=json.dumps(changes, ensure_ascii=False)
        )
        # если AJAX — отдаём JSON
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'redirect_url': str(self.success_url)
            })
        return response

    def form_invalid(self, form):
        # при AJAX — возвращаем ошибки
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'errors': form.errors})
        return super().form_invalid(form)




# AJAX-обработчик для динамической загрузки названий объектов по выбранному типу.
@login_required
def get_object_names(request):
    object_type_id = request.GET.get('object_type')
    data = []
    if object_type_id:
        objects = Object.objects.filter(object_type_id=object_type_id).values('id', 'object_name')
        data = list(objects)
    return JsonResponse(data, safe=False)


# AJAX-обработчик для массового обновления статуса заявок.
@require_POST
@login_required
def mass_update_status(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("Доступ запрещен")

    try:
        data = json.loads(request.body)
        request_ids = data.get('request_ids', [])
        new_status = data.get('new_status')

        if not request_ids or new_status not in ['Новая', 'В работе', 'Выполнена']:
            return JsonResponse({'success': False, 'error': 'Неверные параметры'})

        # Получаем старые статусы для «Undo»
        qs = FlightRequest.objects.filter(id__in=request_ids)
        old_statuses = {str(fr.id): fr.status for fr in qs}

        # Создаём запись UndoAction
        undo = UndoAction.objects.create(
            user=request.user,
            action_type='mass_status',
            payload={
                'request_ids': request_ids,
                'old_statuses': old_statuses
            }
        )

        # Сама массовая смена статуса + история
        for fr in qs:
            old = fr.status
            fr.status = new_status
            fr.save(update_fields=['status'])
            RequestHistory.objects.create(
                flight_request=fr,
                changed_by=request.user,
                changes=json.dumps({'status': [old, new_status]}, ensure_ascii=False)
            )

        # Возвращаем action_id для последующей отмены
        return JsonResponse({'success': True, 'action_id': undo.id})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

import os
import uuid
from django.conf import settings
from django.core.files.storage import default_storage
# Папка «корзины» для временного хранения файлов
TRASH_DIR = os.path.join(settings.MEDIA_ROOT, 'flight_results', '.trash')

# AJAX-обработчик для массового удаления заявок.
@require_POST
@login_required
def mass_delete_requests(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("Доступ запрещен")

    try:
        data = json.loads(request.body)
        request_ids = data.get('request_ids', [])
        if not request_ids:
            return JsonResponse({'success': False, 'error': 'Нет заявок для удаления'})

        # Получаем заявки вместе с их файлами
        qs = FlightRequest.objects.filter(id__in=request_ids).prefetch_related('result_files')

        requests_data = []
        for fr in qs:
            files_data = []
            for rf in fr.result_files.all():
                file_name = rf.file.name if rf.file else None
                view_link = rf.view_link
                file_size = rf.file_size

                file_entry = {
                    'id': rf.id,
                    'result_type': rf.result_type,
                    'file_name': file_name,
                    'view_link': view_link,
                    'file_size': file_size,
                }

                # Перемещаем сам файл в "корзину"
                if file_name:
                    orig_path = file_name
                    trash_subdir = os.path.join(TRASH_DIR, str(uuid.uuid4()))
                    trash_path = os.path.join(trash_subdir, os.path.basename(orig_path))

                    # Копируем файл в корзину
                    with default_storage.open(orig_path, 'rb') as src:
                        default_storage.save(trash_path, src)
                    # Удаляем оригинал
                    default_storage.delete(orig_path)

                    file_entry['trash_path'] = trash_path

                files_data.append(file_entry)

            # Собираем данные по самой заявке
            requests_data.append({
                'id': fr.id,
                'user_id': fr.user_id,
                'username': fr.username,
                'piket_from': fr.piket_from,
                'piket_to': fr.piket_to,
                'shoot_date_from': fr.shoot_date_from.isoformat() if fr.shoot_date_from else None,
                'shoot_date_to': fr.shoot_date_to.isoformat() if fr.shoot_date_to else None,
                'note': fr.note,
                'kml_file_id': fr.kml_file_id,
                'status': fr.status,
                'created_at': fr.created_at.isoformat(),
                'orthophoto': fr.orthophoto,
                'laser': fr.laser,
                'panorama': fr.panorama,
                'overview': fr.overview,
                'object_type_id': fr.object_type_id,
                'object_name_id': fr.object_name_id,
                'result_files': files_data,
            })

        # Создаём запись для Undo
        undo = UndoAction.objects.create(
            user=request.user,
            action_type='mass_delete',
            payload={'requests': requests_data}
        )

        # Фактическое удаление заявок и связанных файлов
        qs.delete()

        return JsonResponse({'success': True, 'action_id': undo.id})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# AJAX-обработчик для отмены массовыхдействий.
@require_POST
@login_required
def undo_action(request, action_id):
    # Получаем UndoAction и проверяем пользователя
    try:
        action = UndoAction.objects.get(pk=action_id, user=request.user)
    except UndoAction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Действие не найдено'})

    # Проверяем время отката (5 секунд)
    if (timezone.now() - action.created_at).total_seconds() > 5:
        action.delete()
        return JsonResponse({'success': False, 'error': 'Время для отмены истекло'})

    payload = action.payload

    if action.action_type == 'mass_status':
        # Восстанавливаем старые статусы
        for req_id, old_status in payload.get('old_statuses', {}).items():
            FlightRequest.objects.filter(id=int(req_id)).update(status=old_status)

    elif action.action_type == 'mass_delete':
        # Восстанавливаем удалённые заявки и их файлы
        for req_data in payload.get('requests', []):
            # 1) Восстанавливаем саму заявку
            fr = FlightRequest.objects.create(
                id=req_data['id'],
                user_id=req_data['user_id'],
                username=req_data['username'],
                piket_from=req_data['piket_from'],
                piket_to=req_data['piket_to'],
                shoot_date_from=(parse_datetime(req_data['shoot_date_from'])
                                  if req_data['shoot_date_from'] else None),
                shoot_date_to=(parse_datetime(req_data['shoot_date_to'])
                                if req_data['shoot_date_to'] else None),
                note=req_data['note'],
                kml_file_id=req_data['kml_file_id'],
                status=req_data['status'],
                created_at=parse_datetime(req_data['created_at']),
                orthophoto=req_data['orthophoto'],
                laser=req_data['laser'],
                panorama=req_data['panorama'],
                overview=req_data['overview'],
                object_type_id=req_data['object_type_id'],
                object_name_id=req_data['object_name_id'],
            )

            # 2) Восстанавливаем все связанные файлы
            for f in req_data.get('result_files', []):
                original_name = f.get('file_name')
                trash_path    = f.get('trash_path')
                view_link     = f.get('view_link')
                file_size     = f.get('file_size')
                file_field    = None

                # Если файл был на диске — копируем из корзины или из оригинального места
                src_path = trash_path or original_name
                if src_path:
                    # Читаем содержимое
                    with default_storage.open(src_path, 'rb') as src:
                        # Сохраняем обратно в flight_results/
                        file_field = default_storage.save(original_name, src)
                    # Удаляем временный файл из корзины
                    if trash_path:
                        default_storage.delete(trash_path)

                # Создаём запись в БД
                FlightResultFile.objects.create(
                    flight_request=fr,
                    result_type=f['result_type'],
                    file=file_field,
                    view_link=view_link,
                    file_size=file_size,
                )

    else:
        return JsonResponse({'success': False, 'error': 'Неподдерживаемый тип действия'})

    # Удаляем запись об откате и возвращаем успех
    action.delete()
    return JsonResponse({'success': True})



import io
from django.template.loader import render_to_string
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame, PageTemplate
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.utils import timezone
from django.utils.formats import date_format as DateFormat
from django.utils.timezone import localtime
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Экспорт заявок в Excel
class ExportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        qs = FlightRequest.objects.all()
        
        # 1.1. Применяем фильтрацию по GET-параметрам
        status = request.GET.get('status')
        object_type = request.GET.get('object_type')
        object_name = request.GET.get('object_name')
        shooting_type = request.GET.get('shooting_type')
        shoot_date_from = request.GET.get('shoot_date_from')
        shoot_date_to = request.GET.get('shoot_date_to')
        
        if status:
            qs = qs.filter(status=status)
        if object_type:
            qs = qs.filter(object_type_id=object_type)
        if object_name:
            qs = qs.filter(object_name_id=object_name)
        if shooting_type:
            qs = qs.filter(**{shooting_type: True})
        if shoot_date_from and shoot_date_to:
            qs = qs.filter(shoot_date_from__lte=shoot_date_to, shoot_date_to__gte=shoot_date_from)
        elif shoot_date_from:
            qs = qs.filter(shoot_date_to__gte=shoot_date_from)
        elif shoot_date_to:
            qs = qs.filter(shoot_date_from__lte=shoot_date_to)
        
        # 1.2. Применяем сортировку
        sort_field = request.GET.get('sort')
        order = request.GET.get('order', 'asc')
        if sort_field == 'shoot_date':
            qs = qs.order_by('-shoot_date_to' if order == 'desc' else 'shoot_date_from')
        else:
            allowed_sort_fields = {
                'status': 'status',
                'id': 'id',
                'object_type': 'object_type__type_name'
            }
            if sort_field in allowed_sort_fields:
                field_name = allowed_sort_fields[sort_field]
                qs = qs.order_by('-' + field_name if order == 'desc' else field_name)
            else:
                qs = qs.order_by('-created_at')
        
        # 2. Формируем HTTP-ответ с типом контента для Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        export_date_str = DateFormat(localtime(timezone.now())).format('Y-m-d_H-i')
        filename = f"Flight_Requests_{export_date_str}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 3. Создаем рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Flight Requests"
        
        # 4. Заголовок файла и дата экспорта
        ws.merge_cells('A1:I1')
        ws['A1'] = "Список заявок"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:I2')
        ws['A2'] = f"Дата выгрузки: {DateFormat(localtime(timezone.now())).format('d.m.Y H:i')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        
        # Начинаем данные с 4 строки
        row_num = 4
        
        # 5. Заголовки столбцов и их ширины
        columns = ['Статус', '№', 'Тип объекта', 'Название объекта', 'Пикеты', 'Дата съемки', 'Тип съемки', 'Примечание', 'Создатель']
        col_widths = [14, 8, 20, 30, 20, 30, 17, 25, 20]  # ширины в единицах Excel (примерные)
        
        # Заполнение заголовков
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = col_widths[col_num - 1]
        
        # 6. Заполнение строк данными
        for req in qs:
            row_num += 1
            # Формируем ячейки для каждого столбца
            # Pickets
            pikets = f"ПК{req.piket_from} – ПК{req.piket_to}" if req.piket_from and req.piket_to else ""
            # Shoot Date
            shoot_date = (f"{req.shoot_date_from.strftime('%d.%m.%Y')} - {req.shoot_date_to.strftime('%d.%m.%Y')}"
                          if req.shoot_date_from and req.shoot_date_to else "")
            # Shooting Type
            shooting_types = []
            if req.orthophoto:
                shooting_types.append("Ортофотоплан")
            if req.laser:
                shooting_types.append("Лазерное сканирование")
            if req.panorama:
                shooting_types.append("Панорама")
            if req.overview:
                shooting_types.append("Обзорные фото")

            shooting_text = ", ".join(shooting_types)
            # Object Name
            try:
                object_name_value = req.object_name.object_name
            except FlightRequest.object_name.RelatedObjectDoesNotExist:
                object_name_value = ""
            # Object Type
            object_type_value = req.object_type.type_name if req.object_type else ""
            
            row_data = [
                req.status,
                req.id,
                object_type_value,
                object_name_value,
                pikets,
                shoot_date,
                shooting_text,
                req.note,
                req.username,
            ]
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
        
        ws.freeze_panes = "A5"
        
        # 7. Сохраняем книгу в HttpResponse
        wb.save(response)
        return response





# Экспорт заявок в PDF
class ExportPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        qs = FlightRequest.objects.all()
        
        # 1.1. Применяем фильтрацию по GET-параметрам
        status = request.GET.get('status')
        object_type = request.GET.get('object_type')
        object_name = request.GET.get('object_name')
        shooting_type = request.GET.get('shooting_type')
        shoot_date_from = request.GET.get('shoot_date_from')
        shoot_date_to = request.GET.get('shoot_date_to')
        
        if status:
            qs = qs.filter(status=status)
        if object_type:
            qs = qs.filter(object_type_id=object_type)
        if object_name:
            qs = qs.filter(object_name_id=object_name)
        if shooting_type:
            qs = qs.filter(**{shooting_type: True})
        if shoot_date_from and shoot_date_to:
            qs = qs.filter(shoot_date_from__lte=shoot_date_to, shoot_date_to__gte=shoot_date_from)
        elif shoot_date_from:
            qs = qs.filter(shoot_date_to__gte=shoot_date_from)
        elif shoot_date_to:
            qs = qs.filter(shoot_date_from__lte=shoot_date_to)
        
        # 1.2. Применяем сортировку
        sort_field = request.GET.get('sort')
        order = request.GET.get('order', 'asc')
        if sort_field == 'shoot_date':
            if order == 'desc':
                qs = qs.order_by('-shoot_date_to')
            else:
                qs = qs.order_by('shoot_date_from')
        else:
            allowed_sort_fields = {
                'status': 'status',
                'id': 'id',
                'object_type': 'object_type__type_name'
            }
            if sort_field in allowed_sort_fields:
                field_name = allowed_sort_fields[sort_field]
                qs = qs.order_by('-' + field_name) if order == 'desc' else qs.order_by(field_name)
            else:
                qs = qs.order_by('-created_at')
        
        # 2. Формируем HttpResponse с типом контента PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f"Flight_Requests_{DateFormat(localtime(timezone.now())).format('d.m.Y_H-i')}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'


        
        # 3. Создаем буфер и документ в альбомной ориентации A4
        buffer = io.BytesIO()
        export_date_str = DateFormat(localtime(timezone.now())).format('Y-m-d_H-i')
        filename = f"Flight_Requests_{export_date_str}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'

        doc_title = f"Flight Requests {export_date_str}"

        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=20,
            rightMargin=20,
            topMargin=15,
            bottomMargin=15,
            title=doc_title
        )

        elements = []
        
        # 4. Регистрация кириллического шрифта (DejaVuSans)
        font_dir = os.path.join(settings.BASE_DIR, 'static', 'fonts')
        font_path_regular = os.path.join(font_dir, 'DejaVuSans.ttf')
        font_path_bold = os.path.join(font_dir, 'DejaVuSans-Bold.ttf')
        if os.path.exists(font_path_regular) and os.path.exists(font_path_bold):
            pdfmetrics.registerFont(TTFont('DejaVuSans', font_path_regular))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', font_path_bold))
            base_font = 'DejaVuSans'
            bold_font = 'DejaVuSans-Bold'
        else:
            base_font = 'Helvetica'
            bold_font = 'Helvetica-Bold'
        
        # 5. Создаем стили с уменьшенным размером шрифта
        styles = getSampleStyleSheet()
        style_title = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontName=bold_font,
            fontSize=14,
            alignment=1
        )
        style_date = ParagraphStyle(
            'Date',
            parent=styles['Normal'],
            fontName=base_font,
            fontSize=10,
            alignment=1
        )
        style_data = ParagraphStyle(
            'Data',
            parent=styles['Normal'],
            fontName=base_font,
            fontSize=9,
            leading=10,
            alignment=0
        )
        style_header = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontName=bold_font,
            fontSize=9,
            leading=10,
            alignment=1
        )


        
        # 6. Заголовок и дата экспорта
        elements.append(Paragraph("Список заявок", style_title))
        elements.append(Spacer(1, 8))
        export_date_str = DateFormat(localtime(timezone.now())).format('d.m.Y H:i')
        elements.append(Paragraph(f"Дата экспорта: {export_date_str}", style_date))
        elements.append(Spacer(1, 16))
        
        # 7. Описание столбцов
        headings = (
            "Статус",
            "№",
            "Тип объекта",
            "Название объекта",
            "Пикеты",
            "Дата съемки",
            "Тип съемки",
            "Примечание",
            "Создатель"
        )
        # Оборачиваем каждую ячейку заголовка в Paragraph для переноса, если необходимо
        table_headings = [Paragraph(h, style_header) for h in headings]
        table_data = [table_headings]
        
        # 8. Формирование строк данных
        for req in qs:
            pikets = f"ПК{req.piket_from} – ПК{req.piket_to}" if req.piket_from and req.piket_to else ""
            shoot_date = (
                f"{req.shoot_date_from.strftime('%d.%m.%Y')} - {req.shoot_date_to.strftime('%d.%m.%Y')}"
                if req.shoot_date_from and req.shoot_date_to else ""
            )
            shooting_types = []
            if req.orthophoto:
                shooting_types.append("Ортофотоплан")
            if req.laser:
                shooting_types.append("Лазерное сканирование")
            if req.panorama:
                shooting_types.append("Панорама")
            if req.overview:
                shooting_types.append("Обзорные фото")
            shooting_text = ", ".join(shooting_types)
            try:
                object_name_value = req.object_name.object_name
            except FlightRequest.object_name.RelatedObjectDoesNotExist:
                object_name_value = ""
            object_type_value = req.object_type.type_name if req.object_type else ""
            
            row = [
                Paragraph(req.status, style_data),
                Paragraph(str(req.id), style_data),
                Paragraph(object_type_value, style_data),
                Paragraph(object_name_value, style_data),
                Paragraph(pikets, style_data),
                Paragraph(shoot_date, style_data),
                Paragraph(shooting_text, style_data),
                Paragraph(req.note if req.note else "", style_data),
                Paragraph(req.username if req.username else "", style_data),
            ]
            table_data.append(row)
        
        # 9. Задаем ширину столбцов так, чтобы таблица влезала на горизонтальном A4
        col_widths = [
            1.0*inch,
            0.5*inch,
            1.2*inch,
            1.7*inch,
            1.15*inch,
            1.3*inch, 
            1.35*inch,
            1.4*inch,
            1.3*inch,
        ]

        # 10. Создаем таблицу
        table = Table(table_data, colWidths=col_widths)
        table_style = TableStyle([
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 1), (-1, 1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ])


        table.setStyle(table_style)
        elements.append(table)
        
        # 11. Настройка шаблона страниц
        frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
        template = PageTemplate(id='template', frames=[frame])
        doc.addPageTemplates([template])
        
        # 12. Формируем PDF
        doc.build(elements)
        pdf_value = buffer.getvalue()
        buffer.close()
        response.write(pdf_value)
        return response



@require_POST
def delete_flight_request(request, pk):
    # Получаем заявку по pk
    flight_request = get_object_or_404(FlightRequest, pk=pk)
    
    # Проверяем права: разрешаем удаление, если пользователь является администратором
    # или если пользователь является создателем заявки
    if not (request.user.is_staff or flight_request.user_id == request.user.id):
        return HttpResponseForbidden("У вас нет прав для удаления этой заявки.")
    
    flight_request.delete()
    return JsonResponse({'success': True})

import os
@require_GET
@login_required
def get_request_results(request):
    """
    Возвращает данные результатов съёмки для заявки.
    Если в заявке установлен тип съемки, но файлов нет – возвращаются заглушки.
    """
    request_id = request.GET.get('request_id')
    if not request_id:
        return HttpResponseBadRequest("request_id не указан")
    
    try:
        flight = FlightRequest.objects.get(id=request_id)
    except FlightRequest.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Заявка не найдена'}, status=404)
    
    data = {}

    # Ортофотоплан
    if flight.orthophoto:
        ortho_files = FlightResultFile.objects.filter(flight_request=flight, result_type='ortho')
        if ortho_files.exists():
            file_obj = ortho_files.first()
            data['orthophoto'] = {
                'id': file_obj.id,
                'download_link': file_obj.file.url if file_obj.file else '#',
                'archive_name': os.path.basename(file_obj.file.name) if file_obj.file else 'Файл не найден'
            }
        else:
            data['orthophoto'] = {
                'download_link': '#',
                'archive_name': 'Файлы для скачивания ещё не добавлены'
            }

    # Лазерное сканирование
    if flight.laser:
        laser_files = FlightResultFile.objects.filter(flight_request=flight, result_type='laser')
        if laser_files.exists():
            file_obj = laser_files.first()
            data['laser'] = {
                'id': file_obj.id,
                'download_link': file_obj.file.url if file_obj.file else '#',
                'archive_name': os.path.basename(file_obj.file.name) if file_obj.file else 'Файл не найден',
                'view_link': file_obj.view_link or '#'
            }
        else:
            data['laser'] = {
                'download_link': '#',
                'archive_name': 'Файлы для скачивания ещё не добавлены',
                'view_link': '#'
            }

    # Панорама
    if flight.panorama:
        panorama_files = FlightResultFile.objects.filter(flight_request=flight, result_type='panorama')
        if panorama_files.exists():
            file_obj = panorama_files.first()
            data['panorama'] = {
                'id': file_obj.id,
                'view_link': file_obj.view_link or '#'
            }
        else:
            data['panorama'] = {
                'view_link': '#'
            }

    # Обзорные фото
    if flight.overview:
        overview_files = FlightResultFile.objects.filter(flight_request=flight, result_type='overview')
        if overview_files.exists():
            file_obj = overview_files.first()
            data['overview'] = {
                'id': file_obj.id,
                'download_link': file_obj.file.url if file_obj.file else '#',
                'archive_name': os.path.basename(file_obj.file.name) if file_obj.file else 'Файл не найден',
                'view_link': file_obj.view_link or '#'
            }
        else:
            data['overview'] = {
                'download_link': '#',
                'archive_name': 'Файлы для скачивания ещё не добавлены',
                'view_link': '#'
            }

    return JsonResponse(data)


@method_decorator(login_required, name='dispatch')
class UploadResultView(View):
    """
    Обрабатывает загрузку результатов съемки.
    """
    def post(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Доступ запрещен'}, status=403)
        
        upload_type = request.POST.get('upload_type')
        request_id = request.POST.get('request_id')
        if not upload_type or not request_id:
            return HttpResponseBadRequest("upload_type и request_id обязательны")
        
        try:
            flight = FlightRequest.objects.get(id=request_id)
        except FlightRequest.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Заявка не найдена'}, status=404)
        
        # Выбираем форму в зависимости от upload_type
        if upload_type == 'ortho':
            form = OrthoUploadForm(request.POST, request.FILES)
        elif upload_type == 'laser':
            form = LaserUploadForm(request.POST, request.FILES)
        elif upload_type == 'panorama':
            form = PanoramaUploadForm(request.POST)
        elif upload_type == 'overview':
            form = OverviewUploadForm(request.POST, request.FILES)
        else:
            return JsonResponse({'success': False, 'error': 'Неверный тип загрузки'}, status=400)

        if form.is_valid():
            if upload_type == 'ortho':
                # Используем новое имя поля "orthoFile"
                temp_file = form.cleaned_data['orthoFile']
                view_link = form.cleaned_data.get('view_link')
                temp_result = TempResultFile.objects.create(
                    uploaded_by=request.user,
                    result_type=upload_type,
                    file=temp_file,
                    view_link=view_link or '',
                    file_size=temp_file.size
                )
                temp_id = temp_result.id
            elif upload_type == 'laser':
                # Используем новое имя поля "laserFile"
                temp_file = form.cleaned_data['laserFile']
                view_link = form.cleaned_data.get('laserViewInput') or None
                temp_result = TempResultFile.objects.create(
                    uploaded_by=request.user,
                    result_type=upload_type,
                    file=temp_file,
                    view_link=view_link,
                    file_size=temp_file.size
                )
                temp_id = temp_result.id
            elif upload_type == 'panorama':
                view_link = form.cleaned_data['panoramaViewInput']
                temp_result = TempResultFile.objects.create(
                    uploaded_by=request.user,
                    result_type=upload_type,
                    file=None,
                    view_link=view_link
                )
                temp_id = temp_result.id
            elif upload_type == 'overview':
                # Получаем единичный файл из поля overviewFiles
                temp_file = request.FILES.get('overviewFiles')
                if not temp_file:
                    return JsonResponse({'success': False, 'error': 'Не выбран файл для загрузки.'}, status=400)
                temp_obj = TempResultFile.objects.create(
                    uploaded_by=request.user,
                    result_type=upload_type,
                    file=temp_file,
                    file_size=temp_file.size
                )
                temp_id = temp_obj.id


            return JsonResponse({
                'success': True,
                'temp_id': temp_id
            })
        else:
            errors = form.errors.as_json()
            return JsonResponse({'success': False, 'error': errors}, status=400)



from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from .models import TempResultFile


@method_decorator(login_required, name='dispatch')
class UploadTempFileView(View):
    """
    Принимает файл, проверяет его (через формы или вручную),
    сохраняет во TempResultFile. Возвращает ID созданного TempResultFile.
    """

    def post(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Доступ запрещен'}, status=403)

        upload_type = request.POST.get('upload_type')  # ortho, laser и т.д.
        if not upload_type:
            return HttpResponseBadRequest("upload_type обязателен")

        if upload_type == 'ortho':
            form = OrthoUploadForm(request.POST, request.FILES)
        elif upload_type == 'laser':
            form = LaserUploadForm(request.POST, request.FILES)
        elif upload_type == 'panorama':
            form = PanoramaUploadForm(request.POST)
        elif upload_type == 'overview':
            form = OverviewUploadForm(request.POST, request.FILES)
        else:
            return JsonResponse({'success': False, 'error': 'Неверный тип загрузки'}, status=400)

        if form.is_valid():
            if upload_type == 'ortho':
                temp_file = form.cleaned_data['orthoFile']
                view_link = form.cleaned_data.get('view_link')
                temp_result = TempResultFile.objects.create(
                    uploaded_by=request.user,
                    result_type=upload_type,
                    file=temp_file,
                    view_link=view_link or '',
                    file_size=temp_file.size
                )
                temp_id = temp_result.id
            elif upload_type == 'laser':
                temp_file = form.cleaned_data['laserFile']
                view_link = form.cleaned_data.get('laserViewInput') or None
                temp_result = TempResultFile.objects.create(
                    uploaded_by=request.user,
                    result_type=upload_type,
                    file=temp_file,
                    view_link=view_link or '',
                    file_size=temp_file.size
                )
                temp_id = temp_result.id
            elif upload_type == 'panorama':
                view_link = form.cleaned_data['panoramaViewInput']
                temp_result = TempResultFile.objects.create(
                    uploaded_by=request.user,
                    result_type=upload_type,
                    file=None,
                    view_link=view_link
                )
                temp_id = temp_result.id
            elif upload_type == 'overview':
                temp_file = request.FILES.get('overviewFiles')
                if not temp_file:
                    return JsonResponse({'success': False, 'error': 'Не выбран файл для загрузки.'}, status=400)
                temp_result = TempResultFile.objects.create(
                    uploaded_by=request.user,
                    result_type=upload_type,
                    file=temp_file,
                    file_size=temp_file.size
                )
                temp_id = temp_result.id


            return JsonResponse({
                'success': True,
                'temp_id': temp_id
            })
        else:
            errors = form.errors.as_json()
            return JsonResponse({'success': False, 'error': errors}, status=400)



@method_decorator(login_required, name='dispatch')
class ConfirmTempFileView(View):
    """
    Принимает temp_id (для одиночных файлов) вместе с request_id и view_link.
    Переносит файл(ы) из TempResultFile -> FlightResultFile.
    Удаляет TempResultFile.
    Для раздела "Обзорные фото" ожидается загрузка только одного архива.
    """
    def post(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Доступ запрещен'}, status=403)

        request_id = request.POST.get('request_id')
        if not request_id:
            return HttpResponseBadRequest("request_id обязательны.")

        try:
            flight = FlightRequest.objects.get(pk=request_id)
        except FlightRequest.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Заявка не найдена'}, status=404)

        temp_id = request.POST.get('temp_id', '').strip()
        view_link_param = request.POST.get('view_link', '').strip()

        # Если temp_id отсутствует, но передана ссылка — обрабатываем только обновление ссылки
        if not temp_id and view_link_param:
            existing_obj = FlightResultFile.objects.filter(
                flight_request=flight,
                result_type='laser'
            ).first()
            if existing_obj:
                existing_obj.view_link = view_link_param
                existing_obj.save()
                return JsonResponse({'success': True, 'message': 'Ссылка успешно обновлена'})
            else:
                final_obj = FlightResultFile.objects.create(
                    flight_request=flight,
                    result_type='laser',
                    view_link=view_link_param
                )
                return JsonResponse({'success': True, 'message': 'Ссылка успешно сохранена (новая запись создана)'})

        # Если temp_id всё ещё отсутствует, выдаём ошибку
        if not temp_id:
            return HttpResponseBadRequest("temp_id обязательны.")

        try:
            temp_file = TempResultFile.objects.get(pk=temp_id, uploaded_by=request.user)
        except TempResultFile.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Temp file не найден'}, status=404)

        # Определяем итоговое значение ссылки:
        # Если значение в POST не пустое, используем его; иначе — устанавливаем None.
        final_view_link = view_link_param if view_link_param != '' else None

        existing_obj = FlightResultFile.objects.filter(
            flight_request=flight,
            result_type=temp_file.result_type
        ).first()

        if existing_obj:
            # Если пользователь не ввёл новую ссылку (final_view_link is None),
            # но в существующей записи уже есть ссылка, сохраняем её.
            if final_view_link is None and existing_obj.view_link:
                final_view_link = existing_obj.view_link
            existing_obj.view_link = final_view_link

            if temp_file.file:
                import os
                filename = os.path.basename(temp_file.file.name)
                if existing_obj.file:
                    existing_obj.file.delete(save=False)
                with temp_file.file.open('rb') as f:
                    existing_obj.file.save(filename, f, save=False)
                existing_obj.file_size = temp_file.file.size
            existing_obj.save()
            final_obj = existing_obj
        else:
            final_obj = FlightResultFile(
                flight_request=flight,
                result_type=temp_file.result_type,
                view_link=final_view_link,
                file_size=temp_file.file_size
            )
            if temp_file.file:
                import os
                filename = os.path.basename(temp_file.file.name)
                with temp_file.file.open('rb') as f:
                    final_obj.file.save(filename, f, save=False)
            final_obj.save()

        if temp_file.file:
            temp_file.file.delete(save=False)
        temp_file.delete()

        return JsonResponse({'success': True, 'message': 'Файл (и/или ссылка) успешно сохранены'})





@method_decorator(login_required, name='dispatch')
class CancelTempFileView(View):
    """
    Удаляет временный файл (при отмене).
    """
    def post(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Доступ запрещен'}, status=403)

        temp_id = request.POST.get('temp_id')
        if not temp_id:
            return HttpResponseBadRequest("temp_id обязателен")

        try:
            temp_file = TempResultFile.objects.get(pk=temp_id, uploaded_by=request.user)
            # Удаляем физический файл из хранилища
            if temp_file.file:
                temp_file.file.delete(save=False)
            temp_file.delete()
            return JsonResponse({'success': True, 'message': 'Временный файл удалён'})
        except TempResultFile.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Temp file не найден'}, status=404)

@method_decorator(login_required, name='dispatch')
class DeleteResultFileView(View):
    """
    Представление для удаления подтверждённого файла (или ссылки)
    из постоянного хранилища.
    """
    def post(self, request, file_id, *args, **kwargs):
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Доступ запрещен'}, status=403)

        element_type = request.POST.get('element_type', 'full')
        
        try:
            file_obj = FlightResultFile.objects.get(pk=file_id)
        except FlightResultFile.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Файл не найден'}, status=404)
        
        # 1. Если удаляем только ссылку для просмотра лазера:
        if element_type == "laser_view":
            file_obj.view_link = None
            file_obj.save()
            # Если в записи после удаления ссылки нет файла, удалим запись полностью.
            if not file_obj.file:
                file_obj.delete()
                return JsonResponse({'success': True, 'message': 'Ссылка удалена, запись пуста – удалена полностью'})
            return JsonResponse({'success': True, 'message': 'Ссылка удалена'})
        
        # 2. Если удаляем файл лазерного сканирования
        if element_type == "laser":
            if file_obj.view_link and file_obj.view_link.strip() != "":
                # Если ссылка сохранена, удаляем только файл
                if file_obj.file:
                    file_obj.file.delete(save=False)
                file_obj.file = None
                file_obj.file_size = None
                file_obj.save()
                return JsonResponse({'success': True, 'message': 'Файл удалён, ссылка сохранена'})
        
        # 3. Для остальных случаев (или если для лазера ссылка отсутствует) – удаляем полностью:
        if file_obj.file:
            file_obj.file.delete(save=False)
        file_obj.delete()
        return JsonResponse({'success': True, 'message': 'Элемент удалён'})